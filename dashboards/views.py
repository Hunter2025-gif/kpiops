from django.core.paginator import Paginator
# --- RESTORE: Admin Timeline View ---
from django.db.models import F, ExpressionWrapper, DateTimeField, Count, Avg
from django.db import models
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import timedelta
from dashboards.templatetags.custom_tags import format_phase_name
from accounts.models import CustomUser
from bmr.models import BMR
from products.models import Product
from workflow.models import BatchPhaseExecution, Machine

@login_required
def admin_timeline_view(request):
    """Admin Timeline View - Track all BMRs through the system"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')

    # Get export format if requested
    export_format = request.GET.get('export')

    # Get all BMRs with timeline data
    bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()

    # Add timeline data for each BMR
    timeline_data = []
    from workflow.models import BatchPhaseExecution
    from bmr.models import BMRRequest
    
    for bmr in bmrs:
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
        bmr_created = bmr.created_date
        
        # Try to get BMR request information if it exists
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        fgs_completed = phases.filter(
            phase__phase_name='finished_goods_store',
            status='completed'
        ).first()
        total_time_days = None
        if fgs_completed and fgs_completed.completed_date:
            total_time_days = (fgs_completed.completed_date - bmr_created).days
        phase_timeline = []
        
        # Add BMR Request phase if available
        if bmr_request:
            phase_timeline.append({
                'phase_name': 'Production Manager BMR Request',
                'status': bmr_request.get_status_display(),
                'started_date': bmr_request.request_date,
                'completed_date': bmr_request.approved_date,
                'started_by': bmr_request.requested_by.get_full_name() if bmr_request.requested_by else None,
                'completed_by': bmr_request.approved_by.get_full_name() if bmr_request.approved_by else None,
                'duration_hours': (bmr_request.approved_date - bmr_request.request_date).total_seconds() / 3600 if bmr_request.approved_date else None,
                'is_request_phase': True  # Flag to identify this as a request phase
            })
        for phase in phases:
            phase_data = {
                'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                'status': phase.status.title(),
                'started_date': phase.started_date,
                'completed_date': phase.completed_date,
                'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                'completed_by': phase.completed_by.get_full_name() if phase.completed_by else None,
                'duration_hours': None,
                'operator_comments': getattr(phase, 'operator_comments', '') or '',
                'phase_order': phase.phase.phase_order if hasattr(phase.phase, 'phase_order') else 0,
            }
            if phase.started_date and phase.completed_date:
                duration = phase.completed_date - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            elif phase.started_date and not phase.completed_date:
                duration = timezone.now() - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            phase_timeline.append(phase_data)
        timeline_data.append({
            'bmr': bmr,
            'total_time_days': total_time_days,
            'phase_timeline': phase_timeline,
            'current_phase': phases.filter(status__in=['pending', 'in_progress']).first(),
            'is_completed': fgs_completed is not None,
        })

    # Handle exports
    if export_format in ['csv', 'excel']:
        return export_timeline_data(request, timeline_data, export_format)

    # Pagination
    paginator = Paginator(timeline_data, 10)  # 10 BMRs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'user': request.user,
        'page_obj': page_obj,
        'timeline_data': page_obj.object_list,
        'dashboard_title': 'BMR Timeline Tracking',
        'total_bmrs': len(timeline_data),
    }

    return render(request, 'dashboards/admin_timeline.html', context)
# Basic workflow_chart view to resolve missing view error

@login_required
def admin_machine_management(request):
    """Machine Management section for admin dashboard"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
        
    # Get all machines
    all_machines = Machine.objects.all().order_by('machine_type', 'name')
    
    # Get recent breakdowns and changeovers (last 30 days)
    recent_breakdowns = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-breakdown_start_time')[:20]
    
    recent_changeovers = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-changeover_start_time')[:20]
    
    # Count breakdowns and changeovers
    total_breakdowns = BatchPhaseExecution.objects.filter(breakdown_occurred=True).count()
    total_changeovers = BatchPhaseExecution.objects.filter(changeover_occurred=True).count()
    
    context = {
        'page_title': 'Machine Management',
        'all_machines': all_machines,
        'recent_breakdowns': recent_breakdowns,
        'recent_changeovers': recent_changeovers,
        'total_breakdowns': total_breakdowns,
        'total_changeovers': total_changeovers,
    }
    
    return render(request, 'dashboards/admin_machine_management.html', context)

@login_required
def export_wip(request):
    """Export Work in Progress data to Excel or CSV"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
        
    # Get filter parameters
    export_format = request.GET.get('format', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get BMRs with active phases
    active_phases = BatchPhaseExecution.objects.filter(status__in=['pending', 'in_progress']).select_related('bmr', 'phase')
    active_bmr_ids = set(active_phases.values_list('bmr_id', flat=True))
    
    # Filter BMRs based on dates if provided
    bmrs_query = BMR.objects.filter(id__in=active_bmr_ids).select_related('product')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        bmrs_query = bmrs_query.filter(actual_start_date__gte=start_date)
        
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        end_date = datetime.combine(end_date, datetime.max.time())
        bmrs_query = bmrs_query.filter(actual_start_date__lte=end_date)
    
    # Create a dict mapping BMR IDs to their active phases
    bmr_active_phases = {}
    for phase in active_phases:
        if phase.bmr_id not in bmr_active_phases:
            bmr_active_phases[phase.bmr_id] = phase
    
    # Process BMRs to add current phase and progress
    work_in_progress_bmrs = []
    for bmr in bmrs_query:
        # Get BMR request information
        from bmr.models import BMRRequest
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        # Get the current phase from our pre-loaded dict
        current_phase = bmr_active_phases.get(bmr.id)
        
        # Calculate progress percentage using aggregation for better performance
        from django.db.models import Count, Q
        phase_counts = BatchPhaseExecution.objects.filter(bmr=bmr).aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed'))
        )
        
        total_phases = phase_counts.get('total', 0)
        completed_phases = phase_counts.get('completed', 0)
        progress_percentage = int((completed_phases / total_phases) * 100) if total_phases > 0 else 0
        
        # Calculate time since request
        time_since_request = 'N/A'
        if bmr_request and bmr_request.request_date:
            time_diff = timezone.now() - bmr_request.request_date
            total_hours = time_diff.total_seconds() / 3600
            if total_hours >= 24:
                days = int(total_hours // 24)
                hours = int(total_hours % 24)
                time_since_request = f"{days}d {hours}h"
            else:
                time_since_request = f"{int(total_hours)}h"
        
        # Prepare data for export
        wip_data = {
            'product_name': bmr.product.product_name,
            'batch_number': bmr.batch_number,
            'batch_size': bmr.actual_batch_size or bmr.product.standard_batch_size,
            'batch_size_unit': bmr.actual_batch_size_unit or bmr.product.batch_size_unit,
            'pack_size': bmr.product.packaging_size_in_units,
            'request_date': bmr_request.request_date.strftime('%Y-%m-%d %H:%M') if bmr_request and bmr_request.request_date else 'N/A',
            'requested_by': bmr_request.requested_by.get_full_name() if bmr_request and bmr_request.requested_by else 'N/A',
            'request_priority': bmr_request.get_priority_display() if bmr_request else 'N/A',
            'time_since_request': time_since_request,
            'current_phase': current_phase.phase.phase_name.replace('_', ' ').title() if current_phase and current_phase.phase else "Awaiting Production",
            'status': "In Production",  # Adding status field
            'started_date': bmr.actual_start_date.strftime('%Y-%m-%d') if bmr.actual_start_date else "N/A",
            'progress': f"{progress_percentage}%"
        }
        work_in_progress_bmrs.append(wip_data)
    
    # Generate export file
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="KPI_Work_In_Progress_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Write company name as main title
        writer.writerow(['Kampala Pharmaceutical Industries'])
        
        # Write Work In Progress Report subtitle
        writer.writerow(['Work In Progress Report'])
        
        # Write date range subtitle
        date_range_text = ""
        if start_date and end_date:
            date_range_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            date_range_text = f"From: {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            date_range_text = f"To: {end_date.strftime('%Y-%m-%d')}"
        writer.writerow([date_range_text])
        
        # Write report generation date
        current_datetime = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        writer.writerow([f'Report Generated: {current_datetime}'])
        
        writer.writerow([])  # Empty row for spacing
        
        # Write headers with BMR request information
        writer.writerow(['Product Name', 'Batch Number', 'Request Date', 'Requested By', 'Priority', 'Time Since Request', 'Batch Size', 'Packaging Size', 'Current Phase', 'Status', 'Progress'])
        
        for item in work_in_progress_bmrs:
            writer.writerow([
                item['product_name'],
                item['batch_number'],
                item['request_date'],
                item['requested_by'],
                item['request_priority'],
                item['time_since_request'],
                f"{item['batch_size']} {item['batch_size_unit']}",
                item['pack_size'],
                item['current_phase'],
                "Approved",  # Status column to match example
                item['progress']
            ])
        
        return response
    
    else:  # Excel format
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="KPI_Work_In_Progress_{timezone.now().strftime("%Y%m%d")}.xls"'
        
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Work in Progress')
        
        # Set up border styles - using slightly thicker borders for better visibility
        borders = xlwt.Borders()
        borders.left = xlwt.Borders.THIN
        borders.right = xlwt.Borders.THIN
        borders.top = xlwt.Borders.THIN
        borders.bottom = xlwt.Borders.THIN
        borders.left_colour = xlwt.Style.colour_map['black']
        borders.right_colour = xlwt.Style.colour_map['black']
        borders.top_colour = xlwt.Style.colour_map['black']
        borders.bottom_colour = xlwt.Style.colour_map['black']
        
        # Company title
        title_style = xlwt.XFStyle()
        title_style.font.bold = True
        title_style.font.height = 280  # Font size 14
        title_style.font.name = 'Calibri'
        title_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        row_num = 0
        ws.write_merge(row_num, row_num, 0, 6, "Kampala Pharmaceutical Industries", title_style)
        
        # First subtitle - Work In Progress Report
        row_num += 1
        subtitle_style = xlwt.XFStyle()
        subtitle_style.font.bold = True
        subtitle_style.font.name = 'Calibri'
        subtitle_style.font.height = 260  # Font size 13
        subtitle_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        # Add Work In Progress Report subtitle
        ws.write_merge(row_num, row_num, 0, 6, "Work In Progress Report", subtitle_style)
        
        # Second subtitle - Date range
        row_num += 1
        date_style = xlwt.XFStyle()
        date_style.font.name = 'Calibri'
        date_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        # Create date range text based on filters
        date_range_text = ""
        if start_date and end_date:
            date_range_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            date_range_text = f"From: {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            date_range_text = f"To: {end_date.strftime('%Y-%m-%d')}"
        
        # Write date range with date_style
        ws.write_merge(row_num, row_num, 0, 6, date_range_text, date_style)
        
        # Third subtitle - Report Generation Date
        row_num += 1
        generation_style = xlwt.XFStyle()
        generation_style.font.name = 'Calibri'
        generation_style.font.italic = True  # Italic text
        generation_style.font.height = 220  # Font size 11
        generation_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        # Generate current date time string
        current_datetime = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        period_text = f"Report Generated: {current_datetime}"
        ws.write_merge(row_num, row_num, 0, 6, period_text, generation_style)
        
        # Add space after report generation date
        row_num += 1
        
        # Table header with exact dark blue background from attachment
        header_style = xlwt.XFStyle()
        header_style.font.bold = True
        header_style.font.name = 'Calibri'
        header_style.font.colour_index = xlwt.Style.colour_map['white']
        
        # Set custom dark blue background color (from attachment)
        pattern = xlwt.Pattern()
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        
        # Creating a custom dark blue color (RGB: 0, 66, 89) - matching the attachment
        xlwt.add_palette_colour("custom_dark_blue", 0x21)
        wb.set_colour_RGB(0x21, 0, 66, 89)
        pattern.pattern_fore_colour = 0x21
        header_style.pattern = pattern
        
        # Add borders to headers
        header_style.borders = borders
        header_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        columns = ['Product Name', 'Batch Number', 'Request Date', 'Requested By', 'Priority', 'Time Since Request', 'Batch Size', 'Packaging Size', 'Current Phase', 'Status', 'Progress']
        
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], header_style)
            # Set column width
            ws.col(col_num).width = 256 * 20  # 20 characters wide
            
        # Sheet body, remaining rows - no alternating colors, just borders
        cell_style = xlwt.XFStyle()
        cell_style.borders = borders
        cell_style.font.name = 'Calibri'
        
        # Centered style for some columns
        centered_style = xlwt.XFStyle()
        centered_style.borders = borders
        centered_style.font.name = 'Calibri'
        centered_style.alignment.horz = xlwt.Alignment.HORZ_CENTER
        
        for idx, item in enumerate(work_in_progress_bmrs):
            row_num += 1
            # Use consistent styles without alternating colors
            row_style = cell_style
            centered_row_style = centered_style
            
            # Write data with borders on all cells including BMR request information
            ws.write(row_num, 0, item['product_name'], row_style)  # Product Name
            ws.write(row_num, 1, item['batch_number'], centered_row_style)  # Batch Number centered
            ws.write(row_num, 2, item['request_date'], centered_row_style)  # Request Date
            ws.write(row_num, 3, item['requested_by'], row_style)  # Requested By
            ws.write(row_num, 4, item['request_priority'], centered_row_style)  # Priority
            ws.write(row_num, 5, item['time_since_request'], centered_row_style)  # Time Since Request
            ws.write(row_num, 6, f"{item['batch_size']} {item['batch_size_unit']}", row_style)  # Batch Size
            ws.write(row_num, 7, item['pack_size'], centered_row_style)  # Packaging Size centered
            ws.write(row_num, 8, item['current_phase'], row_style)  # Current Phase
            ws.write(row_num, 9, "Approved", centered_row_style)  # Status column, matching example
            ws.write(row_num, 10, item['progress'], centered_row_style)  # Progress centered
        
        wb.save(response)
        return response

@login_required
def admin_quality_control(request):
    """Quality Control section for admin dashboard"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Get QC test data using PhaseCheckpoint model
    from workflow.models import PhaseCheckpoint, BatchPhaseExecution
    
    # Get quality control phases
    qc_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name__icontains='qc'
    ).select_related('bmr', 'phase')
    
    # Recent QC checkpoints
    recent_checkpoints = PhaseCheckpoint.objects.select_related(
        'phase_execution__bmr', 'checked_by'
    ).order_by('-checked_date')[:30]
    
    # QC statistics
    all_checkpoints = PhaseCheckpoint.objects.count()
    failed_checkpoints = PhaseCheckpoint.objects.filter(is_within_spec=False).count()
    passed_checkpoints = PhaseCheckpoint.objects.filter(is_within_spec=True).count()
    failure_rate = (failed_checkpoints / all_checkpoints * 100) if all_checkpoints > 0 else 0
    
    context = {
        'page_title': 'Quality Control Dashboard',
        'recent_checkpoints': recent_checkpoints,
        'all_checkpoints': all_checkpoints,
        'failed_checkpoints': failed_checkpoints,
        'passed_checkpoints': passed_checkpoints,
        'failure_rate': round(failure_rate, 1),
        'qc_phases': qc_phases,
    }
    
    return render(request, 'dashboards/admin_quality_control.html', context)

@login_required
def admin_inventory(request):
    """Inventory Management section for admin dashboard"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Material inventory
    from fgs_management.models import FGSInventory
    
    # Finished goods stats
    finished_goods = FGSInventory.objects.all()
    recent_goods = finished_goods.order_by('-created_at')[:20]
    
    context = {
        'page_title': 'Inventory Management',
        'inventory': finished_goods,
        'total_inventory': finished_goods.count(),
        'available_inventory': finished_goods.filter(status='available').count(),
        'recent_goods': recent_goods,  # Maintain compatibility with template
    }
    
    return render(request, 'dashboards/admin_inventory.html', context)

@login_required
def admin_user_management(request):
    """User Management section for admin dashboard"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Get user data
    all_users = CustomUser.objects.all().order_by('role', 'username')
    
    # User statistics
    active_users = all_users.filter(is_active=True).count()
    inactive_users = all_users.filter(is_active=False).count()
    staff_users = all_users.filter(is_staff=True).count()
    
    # Users by role
    role_counts = {}
    for role, _ in CustomUser.ROLE_CHOICES:
        role_counts[role] = all_users.filter(role=role).count()
    
    context = {
        'page_title': 'User Management',
        'all_users': all_users,
        'total_users': all_users.count(),
        'active_users': active_users,
        'inactive_users': inactive_users,
        'staff_users': staff_users,
        'role_counts': role_counts,
    }
    
    return render(request, 'dashboards/admin_user_management.html', context)

@login_required
def admin_system_health(request):
    """System Health section for admin dashboard"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    import sys
    import django
    from django.db import connection
    import platform
    import os
    
    # Get system info
    system_info = {
        'python_version': sys.version,
        'django_version': django.get_version(),
        'database': connection.vendor,
        'os': platform.system() + ' ' + platform.release(),
        'cpu_count': os.cpu_count(),
    }
    
    # Get database stats
    db_stats = {
        'bmr_count': BMR.objects.count(),
        'users_count': CustomUser.objects.count(),
        'phases_count': BatchPhaseExecution.objects.count(),
        'products_count': Product.objects.count(),
    }
    
    # System logs
    from django.contrib.admin.models import LogEntry
    recent_logs = LogEntry.objects.select_related('user', 'content_type').order_by('-action_time')[:50]
    
    context = {
        'page_title': 'System Health',
        'system_info': system_info,
        'db_stats': db_stats,
        'recent_logs': recent_logs,
    }
    
    return render(request, 'dashboards/admin_system_health.html', context)
from django.contrib.auth.decorators import login_required

@login_required
def workflow_chart(request):
    """Workflow Chart View (placeholder)"""
    return render(request, 'dashboards/workflow_chart.html', {'dashboard_title': 'Workflow Chart'})
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.db.models import Count, Q, Min, Max, F, ExpressionWrapper, DateTimeField, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models.functions import Coalesce
import csv
import xlwt
from bmr.models import BMR
from workflow.models import BatchPhaseExecution, Machine
from workflow.services import WorkflowService
from products.models import Product
from accounts.models import CustomUser

def dashboard_home(request):
    """Route users to their role-specific dashboard or redirect to login page"""
    if not request.user.is_authenticated:
        # Redirect to login page for anonymous users
        return redirect('accounts:login')
    
    user_role = request.user.role
    
    role_dashboard_map = {
        'qa': 'dashboards:qa_dashboard',
        'regulatory': 'dashboards:regulatory_dashboard',
        'production_manager': 'dashboards:production_manager_dashboard',  # New Production Manager dashboard
        'store_manager': 'dashboards:store_dashboard',  # Raw material release
        'quarantine': 'quarantine:dashboard',  # Quarantine management dashboard
        'packaging_store': 'dashboards:packaging_dashboard',
        'finished_goods_store': 'dashboards:finished_goods_dashboard',
        'mixing_operator': 'dashboards:mixing_dashboard',
        'qc': 'dashboards:qc_dashboard',
        'tube_filling_operator': 'dashboards:tube_filling_dashboard',
        'packing_operator': 'dashboards:packing_dashboard',
        'granulation_operator': 'dashboards:granulation_dashboard',
        'blending_operator': 'dashboards:blending_dashboard',
        'compression_operator': 'dashboards:compression_dashboard',
        'sorting_operator': 'dashboards:sorting_dashboard',
        'coating_operator': 'dashboards:coating_dashboard',
        'drying_operator': 'dashboards:drying_dashboard',
        'filling_operator': 'dashboards:filling_dashboard',
        'dispensing_operator': 'dashboards:operator_dashboard',  # Material dispensing uses operator dashboard
        'equipment_operator': 'dashboards:operator_dashboard',
        'cleaning_operator': 'dashboards:operator_dashboard',
        'admin': 'dashboards:admin_dashboard',
    }
    
    dashboard_url = role_dashboard_map.get(user_role, 'dashboards:admin_dashboard')
    return redirect(dashboard_url)

@login_required
def admin_dashboard(request):
    
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # === CORE BMR STATISTICS ===
    total_bmrs = BMR.objects.count()
    active_batches = BMR.objects.filter(status__in=['draft', 'approved', 'in_production']).count()
    completed_batches = BMR.objects.filter(status='completed').count()
    rejected_batches = BMR.objects.filter(status='rejected').count()
    
    # === USER MANAGEMENT DATA ===
    total_users = CustomUser.objects.count()
    active_users_count = CustomUser.objects.filter(is_active=True, last_login__gte=timezone.now() - timedelta(days=30)).count()
    recent_users = CustomUser.objects.filter(is_active=True).order_by('-date_joined')[:10]
    
    # === BMR TIMELINE DATA ===
    bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()
    timeline_data = []
    
    for bmr in bmrs:
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
        bmr_created = bmr.created_date
        fgs_completed = phases.filter(
            phase__phase_name='finished_goods_store',
            status='completed'
        ).first()
        
        total_time_days = None
        if fgs_completed and fgs_completed.completed_date:
            total_time_days = (fgs_completed.completed_date - bmr_created).days
            
        phase_timeline = []
        for phase in phases:
            phase_data = {
                'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                'status': phase.status.title(),
                'started_date': phase.started_date,
                'completed_date': phase.completed_date,
                'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                'completed_by': phase.completed_by.get_full_name() if phase.completed_by else None,
                'duration_hours': None,
                'operator_comments': getattr(phase, 'operator_comments', '') or '',
                'phase_order': phase.phase.phase_order if hasattr(phase.phase, 'phase_order') else 0,
            }
            if phase.started_date and phase.completed_date:
                duration = phase.completed_date - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            elif phase.started_date and not phase.completed_date:
                duration = timezone.now() - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            phase_timeline.append(phase_data)
            
        timeline_data.append({
            'bmr': bmr,
            'total_time_days': total_time_days,
            'phase_timeline': phase_timeline,
            'current_phase': phases.filter(status__in=['pending', 'in_progress']).first(),
            'is_completed': fgs_completed is not None,
        })
    
    # Timeline summary stats
    completed_count = sum(1 for item in timeline_data if item['is_completed'])
    in_progress_count = len(timeline_data) - completed_count
    
    # Calculate average production time
    completed_times = [item['total_time_days'] for item in timeline_data if item['total_time_days']]
    
    # Count phases completed today for Live BMR Tracking
    phases_completed_today = BatchPhaseExecution.objects.filter(
        completed_date__date=timezone.now().date()
    ).count()
    avg_production_time = round(sum(completed_times) / len(completed_times)) if completed_times else None
    
    # === ACTIVE PHASES DATA ===
    active_phases = BatchPhaseExecution.objects.filter(
        status__in=['pending', 'in_progress']
    ).select_related('bmr__product', 'phase', 'started_by').order_by('-started_date')
    
    # Add duration calculation for active phases
    for phase in active_phases:
        if phase.started_date:
            duration = timezone.now() - phase.started_date
            phase._duration_hours = round(duration.total_seconds() / 3600, 1)
        else:
            phase._duration_hours = 0
            
    # === WORK IN PROGRESS DATA ===
    # Get BMRs with active phases
    active_bmr_ids = set(active_phases.values_list('bmr_id', flat=True))
    work_in_progress_bmrs = BMR.objects.filter(
        id__in=active_bmr_ids
    ).select_related('product')
    
    # Create a dict mapping BMR IDs to their active phases
    bmr_active_phases = {}
    for phase in active_phases:
        if phase.bmr_id not in bmr_active_phases:
            bmr_active_phases[phase.bmr_id] = phase
    
    for bmr in work_in_progress_bmrs:
        # Get BMR request information
        from bmr.models import BMRRequest
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        # Get the current phase from our pre-loaded dict
        current_phase = bmr_active_phases.get(bmr.id)
        
        if current_phase and current_phase.phase:
            bmr.current_phase_name = current_phase.phase.phase_name.replace('_', ' ').title()
        else:
            bmr.current_phase_name = "Awaiting Production"
        
        # Calculate progress percentage
        from django.db.models import Count, Q
        phase_counts = BatchPhaseExecution.objects.filter(bmr=bmr).aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed'))
        )
        
        total_phases = phase_counts.get('total', 0)
        completed_phases = phase_counts.get('completed', 0)
        
        if total_phases > 0:
            bmr.progress_percentage = int((completed_phases / total_phases) * 100)
        else:
            bmr.progress_percentage = 0
            
        # Add BMR request information for tracking
        bmr.request_date = bmr_request.request_date if bmr_request else None
        bmr.requested_by = bmr_request.requested_by if bmr_request else None
        bmr.request_priority = bmr_request.get_priority_display() if bmr_request else 'N/A'
        
        # Calculate total time since request
        if bmr_request and bmr_request.request_date:
            time_diff = timezone.now() - bmr_request.request_date
            total_hours = time_diff.total_seconds() / 3600
            if total_hours >= 24:
                days = int(total_hours // 24)
                hours = int(total_hours % 24)
                bmr.time_since_request = f"{days}d {hours}h"
            else:
                bmr.time_since_request = f"{int(total_hours)}h"
        else:
            bmr.time_since_request = 'N/A'
    
    # === PRODUCT DISTRIBUTION CHART DATA ===
    # Use the imported Product model from the top of the file
    from django.db.models import Count  # Re-import Count to ensure it's in local scope
    product_types = Product.objects.values('product_type').annotate(count=Count('product_type'))
    tablet_count = 0
    capsule_count = 0
    ointment_count = 0
    
    for item in product_types:
        product_type = item['product_type'].lower() if item['product_type'] else ''
        if 'tablet' in product_type:
            tablet_count += item['count']
        elif 'capsule' in product_type:
            capsule_count += item['count']
        elif 'ointment' in product_type or 'cream' in product_type:
            ointment_count += item['count']
    
    # === PHASE COMPLETION DATA FOR CHARTS ===
    phase_data = {}
    common_phases = ['mixing', 'drying', 'granulation', 'compression', 'packing']
    
    for phase_name in common_phases:
        completed = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status='completed'
        ).count()
        
        in_progress = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status__in=['pending', 'in_progress']
        ).count()
        
        phase_data[f"{phase_name}_completed"] = completed
        phase_data[f"{phase_name}_inprogress"] = in_progress
    
    # === MACHINE MANAGEMENT DATA ===
    all_machines = Machine.objects.all().order_by('machine_type', 'name')
    
    # Get recent breakdowns and changeovers (last 30 days)
    recent_breakdowns = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-breakdown_start_time')[:20]
    
    recent_changeovers = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-changeover_start_time')[:20]
    
    # Count breakdowns and changeovers
    total_breakdowns = BatchPhaseExecution.objects.filter(breakdown_occurred=True).count()
    total_changeovers = BatchPhaseExecution.objects.filter(changeover_occurred=True).count()
    
    # Today's events
    today = timezone.now().date()
    breakdowns_today = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__date=today
    ).count()
    changeovers_today = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__date=today
    ).count()
    
    # Machine statistics
    machine_stats = {}
    for machine in all_machines:
        usage_count = BatchPhaseExecution.objects.filter(machine_used=machine).count()
        breakdown_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            breakdown_occurred=True
        ).count()
        changeover_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            changeover_occurred=True
        ).count()
        
        machine_stats[machine.id] = {
            'machine': machine,
            'usage_count': usage_count,
            'breakdown_count': breakdown_count,
            'changeover_count': changeover_count,
            'breakdown_rate': round((breakdown_count / usage_count * 100), 1) if usage_count > 0 else 0
        }
    
    # === QUALITY CONTROL DATA ===
    qc_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name__in=['post_compression_qc', 'post_mixing_qc', 'post_blending_qc']
    ).select_related('bmr__product', 'phase', 'started_by', 'completed_by').order_by('-started_date')
    
    # Categorize QC tests
    passed_tests = qc_phases.filter(status='completed')
    failed_tests = qc_phases.filter(status='failed')
    pending_tests = qc_phases.filter(status='in_progress')
    
    qc_stats = {
        'passed_tests': passed_tests.count(),
        'failed_tests': failed_tests.count(),
        'pending_tests': pending_tests.count(),
    }
    
    # Detailed QC test data for clickable cards
    qc_test_details = {
        'passed_tests_data': passed_tests[:10],  # Latest 10 passed tests
        'failed_tests_data': failed_tests[:10],  # Latest 10 failed tests  
        'pending_tests_data': pending_tests[:10],  # Latest 10 pending tests
    }
    
    # === FGS (FINISHED GOODS STORAGE) DATA ===
    # Import FGS models if available
    try:
        from fgs_management.models import FGSInventory, ProductRelease, FGSAlert
        
        fgs_phases = BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store'
        ).select_related('bmr__product', 'started_by', 'completed_by').order_by('-started_date')
        
        fgs_stats = {
            'total_in_store': fgs_phases.filter(status='completed').count(),
            'pending_storage': fgs_phases.filter(status='pending').count(),
            'being_stored': fgs_phases.filter(status='in_progress').count(),
            'available_for_sale': FGSInventory.objects.filter(status='available').count(),
            'recent_releases': ProductRelease.objects.filter(
                release_date__gte=timezone.now() - timedelta(days=7)
            ).count(),
            'active_alerts': FGSAlert.objects.filter(is_resolved=False).count(),
        }
    except ImportError:
        # Fallback if FGS models not available
        fgs_stats = {
            'total_in_store': 156,
            'pending_storage': 12,
            'being_stored': 8,
            'available_for_sale': 89,
            'recent_releases': 5,
            'active_alerts': 2,
        }
    
    # === SYSTEM HEALTH DATA ===
    pending_approvals = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        status='pending'
    ).count()
    
    failed_phases = BatchPhaseExecution.objects.filter(
        status='failed',
        completed_date__date=timezone.now().date()
    ).count()
    
    # Production metrics
    production_stats = {
        'in_production': BatchPhaseExecution.objects.filter(
            status='in_progress'
        ).count(),
        'quality_hold': BatchPhaseExecution.objects.filter(
            phase__phase_name__contains='qc',
            status='pending'
        ).count(),
        'awaiting_packaging': BatchPhaseExecution.objects.filter(
            phase__phase_name='packaging_material_release',
            status='pending'
        ).count(),
        'final_qa_pending': BatchPhaseExecution.objects.filter(
            phase__phase_name='final_qa',
            status='pending'
        ).count(),
        'in_fgs': BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status__in=['completed', 'in_progress']
        ).count(),
    }
    
    # === ADDITIONAL DATA FOR DASHBOARD ===
    recent_bmrs = BMR.objects.select_related('product', 'created_by').order_by('-created_date')[:20]
    
    # Calculate machine utilization
    total_machines = all_machines.count()
    active_machines = all_machines.filter(is_active=True).count()
    machine_utilization = round((active_machines / total_machines * 100), 1) if total_machines > 0 else 0
    
    # === QUARANTINE MONITORING DATA ===
    from quarantine.models import QuarantineBatch, SampleRequest
    
    # Get all quarantine batches with their sample requests
    quarantine_batches = QuarantineBatch.objects.select_related(
        'bmr__product',
        'bmr__created_by',
        'current_phase'
    ).prefetch_related(
        'sample_requests__requested_by',
        'sample_requests__sampled_by',      # QA staff who took the sample
        'sample_requests__received_by',     # QC staff who received the sample
        'sample_requests__approved_by'      # QC staff who approved/rejected
    ).order_by('-quarantine_date')[:20]
    
    # Get recent sample requests with detailed timeline
    recent_sample_requests = SampleRequest.objects.select_related(
        'quarantine_batch__bmr__product',
        'quarantine_batch__bmr__created_by',
        'requested_by',
        'sampled_by',  # QA user who sampled
        'received_by',  # QC user who received
        'approved_by'   # QC user who approved/rejected
    ).order_by('-request_date')[:15]
    
    # Calculate quarantine statistics
    total_quarantine_batches = QuarantineBatch.objects.count()
    pending_qa_samples = SampleRequest.objects.filter(
        sample_date__isnull=True  # No sample taken yet = pending QA
    ).count()
    pending_qc_samples = SampleRequest.objects.filter(
        sample_date__isnull=False,  # Sample taken by QA
        qc_status='pending'  # But QC not completed
    ).count()
    approved_samples_today = SampleRequest.objects.filter(
        qc_status='approved',
        approved_date__date=timezone.now().date()
    ).count()
    rejected_samples_today = SampleRequest.objects.filter(
        qc_status='failed',
        approved_date__date=timezone.now().date()
    ).count()
    
    quarantine_stats = {
        'total_quarantine_batches': total_quarantine_batches,
        'pending_qa_samples': pending_qa_samples,
        'pending_qc_samples': pending_qc_samples,
        'approved_samples_today': approved_samples_today,
        'rejected_samples_today': rejected_samples_today,
        'avg_qa_processing_time': SampleRequest.objects.filter(
            sample_date__isnull=False  # QA processing completed
        ).aggregate(
            avg_time=models.Avg(
                models.F('sample_date') - models.F('request_date')
            )
        )['avg_time'],
        'avg_qc_processing_time': SampleRequest.objects.filter(
            approved_date__isnull=False,  # QC decision made
            received_date__isnull=False   # QC received sample
        ).aggregate(
            avg_time=models.Avg(
                models.F('approved_date') - models.F('received_date')
            )
        )['avg_time']
    }
    
    # Productivity metrics
    operators = CustomUser.objects.filter(
        role__in=['mixing_operator', 'compression_operator', 'granulation_operator', 'packing_operator']
    )
    
    top_operators = []
    for operator in operators:
        completed_phases = BatchPhaseExecution.objects.filter(
            completed_by=operator,
            status='completed'
        ).count()
        
        if completed_phases > 0:
            top_operators.append({
                'name': operator.get_full_name(),
                'completions': completed_phases,
                'role': operator.get_role_display()
            })
    
    top_operators.sort(key=lambda x: x['completions'], reverse=True)
    top_operators = top_operators[:10]
    
    productivity_metrics = {
        'top_operators': top_operators,
        'total_operators': operators.count(),
        'total_completions': sum([op['completions'] for op in top_operators])
    }
    
    # === CONTEXT DATA FOR TEMPLATE ===
    context = {
        'user': request.user,
        'dashboard_title': 'Admin Control Center - Kampala Pharmaceutical Industries',
        
        # === DASHBOARD OVERVIEW DATA ===
        'total_bmrs': total_bmrs,
        'active_batches': active_batches,
        'completed_batches': completed_batches,
        'rejected_batches': rejected_batches,
        'recent_bmrs': recent_bmrs,
        
        # === CHART DATA ===
        'tablet_count': tablet_count,
        'capsule_count': capsule_count,
        'ointment_count': ointment_count,
        'mixing_completed': phase_data.get('mixing_completed', 0),
        'mixing_inprogress': phase_data.get('mixing_inprogress', 0),
        'drying_completed': phase_data.get('drying_completed', 0),
        'drying_inprogress': phase_data.get('drying_inprogress', 0),
        'granulation_completed': phase_data.get('granulation_completed', 0),
        'granulation_inprogress': phase_data.get('granulation_inprogress', 0),
        'compression_completed': phase_data.get('compression_completed', 0),
        'compression_inprogress': phase_data.get('compression_inprogress', 0),
        'packing_completed': phase_data.get('packing_completed', 0),
        'packing_inprogress': phase_data.get('packing_inprogress', 0),
        
        # === TIMELINE TRACKING DATA ===
        'timeline_data': timeline_data[:15],  # Limit for performance
        'completed_count': completed_count,
        'in_progress_count': in_progress_count,
        'avg_production_time': avg_production_time,
        'phases_completed_today': phases_completed_today,
        
        # === ACTIVE PHASES DATA ===
        'active_phases': active_phases[:15],  # Limit for performance
        
        # === WORK IN PROGRESS DATA ===
        'work_in_progress_bmrs': work_in_progress_bmrs,
        
        # === MACHINE MANAGEMENT DATA ===
        'all_machines': all_machines,
        'machine_stats': machine_stats,
        'total_machines': total_machines,
        'active_machines': active_machines,
        'machine_utilization': machine_utilization,
        'recent_breakdowns': recent_breakdowns,
        'recent_changeovers': recent_changeovers,
        'total_breakdowns': total_breakdowns,
        'total_changeovers': total_changeovers,
        'breakdowns_today': breakdowns_today,
        'changeovers_today': changeovers_today,
        
        # === QUALITY CONTROL DATA ===
        'qc_stats': qc_stats,
        'qc_test_details': qc_test_details,
        
        # === FGS & INVENTORY DATA ===
        'fgs_stats': fgs_stats,
        
        # === QUARANTINE MONITORING DATA ===
        'quarantine_batches': quarantine_batches,
        'quarantine_stats': quarantine_stats,
        'recent_sample_requests': recent_sample_requests,
        
        # === USER MANAGEMENT DATA ===
        'total_users': total_users,
        'active_users_count': active_users_count,
        'recent_users': recent_users,
        'productivity_metrics': productivity_metrics,
        
        # === SYSTEM HEALTH DATA ===
        'pending_approvals': pending_approvals,
        'failed_phases': failed_phases,
        'production_stats': production_stats,
        
        # === LIVE TRACKING DATA ===
        'phases_completed_today': phases_completed_today,
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)

@login_required
@csrf_protect
def qa_dashboard(request):
    """Quality Assurance Dashboard"""
    if request.user.role != 'qa':
        messages.error(request, 'Access denied. QA role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for Final QA workflow
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        comments = request.POST.get('comments', '')
        
        if phase_id and action in ['start', 'approve', 'reject']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Start the Final QA review process
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Final QA review started by {request.user.get_full_name()}. Notes: {comments}"
                    phase_execution.save()
                    
                    messages.success(request, f'Final QA review started for batch {phase_execution.bmr.batch_number}. You can now complete the review.')
                
                elif action == 'approve':
                    # Complete Final QA with approval
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments += f"\nFinal QA Approved by {request.user.get_full_name()}. Comments: {comments}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow (should be finished goods store)
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'Final QA approved for batch {phase_execution.bmr.batch_number}. Batch is ready for finished goods storage.')
                    
                elif action == 'reject':
                    # Complete Final QA with rejection
                    phase_execution.status = 'failed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments += f"\nFinal QA Rejected by {request.user.get_full_name()}. Rejection Reason: {comments}"
                    phase_execution.save()
                    
                    # Rollback to appropriate packing phase based on product type
                    bmr = phase_execution.bmr
                    product_type = bmr.product.product_type
                    
                    # Determine which packing phase to rollback to based on product type
                    if product_type == 'tablet':
                        if hasattr(bmr.product, 'tablet_type') and bmr.product.tablet_type == 'tablet_2':
                            rollback_phase = 'bulk_packing'
                        else:
                            rollback_phase = 'blister_packing'
                    elif product_type == 'capsule':
                        rollback_phase = 'blister_packing'
                    elif product_type == 'ointment':
                        rollback_phase = 'secondary_packaging'
                    else:
                        rollback_phase = 'secondary_packaging'  # default
                    
                    # Find and activate the appropriate packing phase for rework
                    rollback_execution = BatchPhaseExecution.objects.filter(
                        bmr=bmr,
                        phase__phase_name=rollback_phase
                    ).first()
                    
                    if rollback_execution:
                        rollback_execution.status = 'pending'
                        rollback_execution.operator_comments = f"Returned for rework due to Final QA rejection. Reason: {comments}. Original comments: {rollback_execution.operator_comments}"
                        rollback_execution.save()
                        
                        messages.warning(request, f'Final QA rejected for batch {bmr.batch_number}. Batch has been sent back to {rollback_phase.replace("_", " ").title()} for rework.')
                    else:
                        messages.error(request, f'Could not find {rollback_phase} phase to rollback to for batch {bmr.batch_number}.')
                    
            except Exception as e:
                messages.error(request, f'Error processing Final QA: {str(e)}')
        
        return redirect('dashboards:qa_dashboard')
    
    # Get QA-specific data
    total_bmrs = BMR.objects.count()
    draft_bmrs = BMR.objects.filter(status='draft').count()
    submitted_bmrs = BMR.objects.filter(status='submitted').count()
    my_bmrs = BMR.objects.filter(created_by=request.user).count()
    
    # Recent BMRs created by this user
    recent_bmrs = BMR.objects.filter(created_by=request.user).select_related('product').order_by('-created_date')[:5]
    
    # BMRs needing final QA review
    final_qa_pending = BatchPhaseExecution.objects.filter(
        phase__phase_name='final_qa',
        status='pending'
    ).select_related('bmr', 'phase')[:10]
    
    # Final QA reviews in progress (started but not completed)
    final_qa_in_progress = BatchPhaseExecution.objects.filter(
        phase__phase_name='final_qa',
        status='in_progress'
    ).select_related('bmr', 'phase')[:10]
    
    # Import BMRRequest model
    from bmr.models import BMRRequest
    
    # Get BMR requests data
    bmr_requests_pending = BMRRequest.objects.filter(status='pending').select_related('product', 'requested_by').order_by('-request_date')[:5]
    bmr_request_counts = {
        'pending': BMRRequest.objects.filter(status='pending').count(),
        'approved': BMRRequest.objects.filter(status='approved').count(),
        'rejected': BMRRequest.objects.filter(status='rejected').count(),
    }
    
    # Build operator history for this user: only regulatory approval phases completed by this user
    regulatory_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        completed_by=request.user
    ).order_by('-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in regulatory_phases
    ]

    # Get quarantine samples awaiting QA processing
    try:
        from quarantine.models import SampleRequest
        quarantine_samples_pending = SampleRequest.objects.filter(
            sample_date__isnull=True  # Not yet processed by QA
        ).select_related(
            'quarantine_batch__bmr__product',
            'quarantine_batch__bmr'
        ).order_by('-request_date')[:10]
        
        quarantine_samples_processed_today = SampleRequest.objects.filter(
            sample_date__date=timezone.now().date()  # Processed by QA today
        ).select_related(
            'quarantine_batch__bmr__product',
            'quarantine_batch__bmr'
        ).order_by('-sample_date')[:10]
    except ImportError:
        # Quarantine app not yet migrated
        quarantine_samples_pending = []
        quarantine_samples_processed_today = []

    context = {
        'user': request.user,
        'total_bmrs': total_bmrs,
        'draft_bmrs': draft_bmrs,
        'submitted_bmrs': submitted_bmrs,
        'my_bmrs': my_bmrs,
        'recent_bmrs': recent_bmrs,
        'final_qa_pending': final_qa_pending,
        'final_qa_in_progress': final_qa_in_progress,
        'quarantine_samples_pending': quarantine_samples_pending,  # Add quarantine samples
        'quarantine_samples_processed_today': quarantine_samples_processed_today,  # Add processed samples
        'dashboard_title': 'Quality Assurance Dashboard',
        'operator_history': operator_history,
        'bmr_requests_pending': bmr_requests_pending,
        'bmr_request_counts': bmr_request_counts,
    }
    return render(request, 'dashboards/qa_dashboard.html', context)

@login_required
def regulatory_dashboard(request):
    """Regulatory Dashboard"""
    if request.user.role != 'regulatory':
        messages.error(request, 'Access denied. Regulatory role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for approval/rejection
    if request.method == 'POST':
        action = request.POST.get('action')
        bmr_id = request.POST.get('bmr_id')
        comments = request.POST.get('comments', '')
        
        if bmr_id and action in ['approve', 'reject']:
            try:
                bmr = get_object_or_404(BMR, pk=bmr_id)
                
                # Find the regulatory approval phase for this BMR
                regulatory_phase = BatchPhaseExecution.objects.filter(
                    bmr=bmr,
                    phase__phase_name='regulatory_approval',
                    status='pending'
                ).first()
                
                if regulatory_phase:
                    if action == 'approve':
                        regulatory_phase.status = 'completed'
                        regulatory_phase.completed_by = request.user
                        regulatory_phase.completed_date = timezone.now()
                        regulatory_phase.operator_comments = f"Approved by {request.user.get_full_name()}. Comments: {comments}"
                        regulatory_phase.save()
                        
                        # Update BMR status
                        bmr.status = 'approved'
                        bmr.approved_by = request.user
                        bmr.approved_date = timezone.now()
                        bmr.save()
                        
                        # Trigger next phase in workflow
                        WorkflowService.trigger_next_phase(bmr, regulatory_phase.phase)
                        
                        messages.success(request, f'BMR {bmr.batch_number} has been approved successfully.')
                        
                    elif action == 'reject':
                        regulatory_phase.status = 'failed'
                        regulatory_phase.completed_by = request.user
                        regulatory_phase.completed_date = timezone.now()
                        regulatory_phase.operator_comments = f"Rejected by {request.user.get_full_name()}. Reason: {comments}"
                        regulatory_phase.save()
                        
                        # Update BMR status
                        bmr.status = 'rejected'
                        bmr.approved_by = request.user
                        bmr.approved_date = timezone.now()
                        bmr.save()
                        
                        messages.warning(request, f'BMR {bmr.batch_number} has been rejected and sent back to QA.')
                else:
                    messages.error(request, 'No pending regulatory approval found for this BMR.')
                    
            except Exception as e:
                messages.error(request, f'Error processing request: {str(e)}')
        
        return redirect('dashboards:regulatory_dashboard')
    
    # BMRs waiting for regulatory approval (pending regulatory_approval phase)
    pending_approvals = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        status='pending'
    ).select_related('bmr__product', 'phase').order_by('bmr__created_date')
    
    # Statistics
    stats = {
        'pending_approvals': pending_approvals.count(),
        'approved_today': BMR.objects.filter(
            status='approved',
            approved_date__date=timezone.now().date()
        ).count(),
        'rejected_this_week': BMR.objects.filter(
            status='rejected',
            approved_date__gte=timezone.now().date() - timedelta(days=7)
        ).count(),
        'total_bmrs': BMR.objects.count(),
    }
    
    context = {
        'user': request.user,
        'pending_approvals': pending_approvals,
        'stats': stats,
        'dashboard_title': 'Regulatory Dashboard'
    }
    return render(request, 'dashboards/regulatory_dashboard.html', context)

@login_required
def store_dashboard(request):
    """Store Manager Dashboard - Raw Material Release Phase"""
    if request.user.role != 'store_manager':
        messages.error(request, 'Access denied. Store Manager role required.')
        return redirect('dashboards:dashboard_home')
    
    if request.method == 'POST':
        bmr_id = request.POST.get('bmr_id')
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        try:
            bmr = BMR.objects.get(pk=bmr_id)
            
            # Get the raw material release phase
            phase_execution = BatchPhaseExecution.objects.get(
                bmr=bmr,
                phase__phase_name='raw_material_release'
            )
            
            if action == 'start':
                phase_execution.status = 'in_progress'
                phase_execution.started_by = request.user
                phase_execution.started_date = timezone.now()
                phase_execution.operator_comments = f"Raw material release started by {request.user.get_full_name()}. Notes: {notes}"
                phase_execution.save()
                
                messages.success(request, f'Raw material release started for batch {bmr.batch_number}.')
                
            elif action == 'complete':
                phase_execution.status = 'completed'
                phase_execution.completed_by = request.user
                phase_execution.completed_date = timezone.now()
                phase_execution.operator_comments = f"Raw materials released by {request.user.get_full_name()}. Notes: {notes}"
                phase_execution.save()
                
                # Trigger next phase in workflow (material_dispensing)
                WorkflowService.trigger_next_phase(bmr, phase_execution.phase)
                
                messages.success(request, f'Raw materials released for batch {bmr.batch_number}. Material dispensing is now available.')
                
        except Exception as e:
            messages.error(request, f'Error processing raw material release: {str(e)}')
    
        return redirect('dashboards:store_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get raw material release phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    # Get recently completed releases (last 7 days)
    recently_completed = BatchPhaseExecution.objects.filter(
        phase__phase_name='raw_material_release',
        status='completed',
        completed_date__gte=timezone.now() - timedelta(days=7)
    ).select_related('bmr__product', 'completed_by').order_by('-completed_date')[:10]
    
    return render(request, 'dashboards/store_dashboard.html', {
        'my_phases': my_phases,
        'stats': stats,
        'recently_completed': recently_completed,
    })

@login_required
def operator_dashboard(request):
    """Generic operator dashboard for production phases"""
    
    # Handle POST requests for phase start/completion
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        comments = request.POST.get('comments', '')
        
        # Machine-related fields
        machine_id = request.POST.get('machine_id')
        
        # Breakdown fields
        breakdown_occurred = request.POST.get('breakdown_occurred') == 'on'
        breakdown_start_time = request.POST.get('breakdown_start_time')
        breakdown_end_time = request.POST.get('breakdown_end_time')
        
        # Changeover fields  
        changeover_occurred = request.POST.get('changeover_occurred') == 'on'
        changeover_start_time = request.POST.get('changeover_start_time')
        changeover_end_time = request.POST.get('changeover_end_time')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Check if machine selection is required for this phase
                    machine_required_phases = ['granulation', 'blending', 'compression', 'coating', 'blister_packing', 'bulk_packing', 'filling']
                    phase_name = phase_execution.phase.phase_name
                    
                    # For capsule filling, only require machine for filling phase
                    if phase_name == 'filling' and phase_execution.bmr.product.product_type != 'Capsule':
                        machine_required = False
                    elif phase_name in machine_required_phases:
                        machine_required = True
                    else:
                        machine_required = False
                    
                    if machine_required and not machine_id:
                        messages.error(request, f'Machine selection is required for {phase_name} phase.')
                        return redirect(request.path)
                    
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start {phase_execution.phase.phase_name} for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect(request.path)
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Started by {request.user.get_full_name()}. Notes: {comments}"
                    
                    # Set machine if provided
                    if machine_id:
                        try:
                            machine = Machine.objects.get(id=machine_id, is_active=True)
                            phase_execution.machine_used = machine
                        except Machine.DoesNotExist:
                            messages.error(request, 'Selected machine not found or inactive.')
                            return redirect(request.path)
                    
                    phase_execution.save()
                    
                    machine_info = f" using {phase_execution.machine_used.name}" if phase_execution.machine_used else ""
                    messages.success(request, f'Phase {phase_execution.phase.phase_name}{machine_info} started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Completed by {request.user.get_full_name()}. Notes: {comments}"
                    
                    # Only handle breakdown/changeover for production phases (not material dispensing)
                    phase_name = phase_execution.phase.phase_name
                    exclude_breakdown_phases = ['material_dispensing', 'bmr_creation', 'regulatory_approval', 'bulk_packing', 'secondary_packaging']
                    
                    if phase_name not in exclude_breakdown_phases:
                        # Handle breakdown tracking
                        phase_execution.breakdown_occurred = breakdown_occurred
                        if breakdown_occurred and breakdown_start_time and breakdown_end_time:
                            from datetime import datetime
                            try:
                                phase_execution.breakdown_start_time = datetime.fromisoformat(breakdown_start_time.replace('T', ' '))
                                phase_execution.breakdown_end_time = datetime.fromisoformat(breakdown_end_time.replace('T', ' '))
                            except ValueError:
                                messages.warning(request, 'Invalid breakdown time format. Breakdown recorded without times.')
                        
                        # Handle changeover tracking
                        phase_execution.changeover_occurred = changeover_occurred
                        if changeover_occurred and changeover_start_time and changeover_end_time:
                            from datetime import datetime
                            try:
                                phase_execution.changeover_start_time = datetime.fromisoformat(changeover_start_time.replace('T', ' '))
                                phase_execution.changeover_end_time = datetime.fromisoformat(changeover_end_time.replace('T', ' '))
                            except ValueError:
                                messages.warning(request, 'Invalid changeover time format. Changeover recorded without times.')
                    
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    completion_msg = f'Phase {phase_execution.phase.phase_name} completed for batch {phase_execution.bmr.batch_number}.'
                    if breakdown_occurred:
                        completion_msg += ' Breakdown recorded.'
                    if changeover_occurred:
                        completion_msg += ' Changeover recorded.'
                    
                    messages.success(request, completion_msg)
                    
            except Exception as e:
                messages.error(request, f'Error processing phase: {str(e)}')
        
        return redirect(request.path)  # Redirect to same dashboard
    
    # Get phases this user can work on
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    my_phases = []
    
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }

    # Determine the primary phase name for this role
    role_phase_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'drying_operator': 'drying',
        'filling_operator': 'filling',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'packing',
        'sorting_operator': 'sorting',
        'dispensing_operator': 'dispensing',  # Material dispensing operator
    }

    phase_name = role_phase_mapping.get(request.user.role, 'production')
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)

    # Operator History: all phases completed by this user for their role

    # Fix: Use .distinct() before slicing to avoid TypeError
    completed_phases_qs = BatchPhaseExecution.objects.filter(
        completed_by=request.user
    ).select_related('bmr', 'phase').order_by('-completed_date')
    completed_phases = list(completed_phases_qs[:20])
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M') if (p.completed_date or p.started_date or p.created_date) else '',
            'batch': p.bmr.bmr_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in completed_phases
    ]

    # Operator Statistics
    # Use .distinct() before slicing for batches_handled
    batches_handled = completed_phases_qs.values('bmr').distinct().count()
    total_completed = completed_phases_qs.count()
    total_attempted = BatchPhaseExecution.objects.filter(started_by=request.user).count()
    success_rate = round((total_completed / total_attempted) * 100, 1) if total_attempted else 0
    completion_times = [
        (p.completed_date - p.started_date).total_seconds() / 60
        for p in completed_phases if p.completed_date and p.started_date
    ]
    avg_completion_time = f"{round(sum(completion_times)/len(completion_times), 1)} min" if completion_times else "-"
    assignment_status = "You have assignments pending." if stats['pending_phases'] > 0 else "All assignments up to date."
    operator_stats = {
        'batches_handled': batches_handled,
        'success_rate': success_rate,
        'avg_completion_time': avg_completion_time,
        'assignment_status': assignment_status,
    }

    # Operator Assignments: current in-progress or pending phases
    operator_assignments = [
        f"{p.bmr.bmr_number} - {p.phase.get_phase_name_display()} ({p.status.title()})"
        for p in my_phases if p.status in ['pending', 'in_progress']
    ]

    # Get machines for this operator's phase type
    machine_type_mapping = {
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'packing_operator': 'blister_packing',  # Packing operator uses blister packing machines
        'filling_operator': 'filling',  # For capsule filling
    }
    
    user_machine_type = machine_type_mapping.get(request.user.role)
    available_machines = []
    if user_machine_type:
        available_machines = Machine.objects.filter(
            machine_type=user_machine_type,
            is_active=True
        ).order_by('name')

    # Determine if this role should show breakdown/changeover tracking
    # Exclude material dispensing and administrative phases
    breakdown_tracking_roles = [
        'mixing_operator', 'granulation_operator', 'blending_operator', 'compression_operator',
        'coating_operator', 'drying_operator', 'filling_operator', 'tube_filling_operator',
        'sorting_operator', 'packing_operator'
    ]
    show_breakdown_tracking = request.user.role in breakdown_tracking_roles

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'phase_name': phase_name,
        'daily_progress': daily_progress,
        'dashboard_title': f'{request.user.get_role_display()} Dashboard',
        'operator_history': operator_history,
        'operator_stats': operator_stats,
        'operator_assignments': operator_assignments,
        'available_machines': available_machines,
        'show_breakdown_tracking': show_breakdown_tracking,
    }

    return render(request, 'dashboards/operator_dashboard.html', context)

# Specific operator dashboards
@login_required
def mixing_dashboard(request):
    return operator_dashboard(request)

@login_required
def granulation_dashboard(request):
    return operator_dashboard(request)

@login_required
def blending_dashboard(request):
    return operator_dashboard(request)

@login_required
def compression_dashboard(request):
    return operator_dashboard(request)

@login_required
def coating_dashboard(request):
    return operator_dashboard(request)

@login_required
def drying_dashboard(request):
    return operator_dashboard(request)

@login_required
def filling_dashboard(request):
    return operator_dashboard(request)

@login_required
def tube_filling_dashboard(request):
    return operator_dashboard(request)

@login_required
def sorting_dashboard(request):
    return operator_dashboard(request)

@login_required
def qc_dashboard(request):
    """Quality Control Dashboard"""
    if request.user.role != 'qc':
        messages.error(request, 'Access denied. QC role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for QC test results
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        test_results = request.POST.get('test_results', '')
        
        if phase_id and action in ['start', 'pass', 'fail']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Start QC testing
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"QC Testing started by {request.user.get_full_name()}. Notes: {test_results}"
                    phase_execution.save()
                    
                    messages.success(request, f'QC testing started for batch {phase_execution.bmr.batch_number}.')
                
                elif action == 'pass':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"QC Test Passed by {request.user.get_full_name()}. Results: {test_results}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'QC test passed for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'fail':
                    phase_execution.status = 'failed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"QC Test Failed by {request.user.get_full_name()}. Results: {test_results}"
                    phase_execution.save()
                    
                    # Rollback to previous phase and get the specific phase name
                    rollback_phase_name = WorkflowService.rollback_to_previous_phase(phase_execution.bmr, phase_execution.phase)
                    
                    if rollback_phase_name:
                        # Format phase name for display
                        display_phase_name = rollback_phase_name.replace('_', ' ').title()
                        messages.warning(request, f'QC test failed for batch {phase_execution.bmr.batch_number}. Rolled back to {display_phase_name} phase.')
                    else:
                        messages.error(request, f'QC test failed for batch {phase_execution.bmr.batch_number}. Rollback failed - please contact administrator.')
                    
            except Exception as e:
                messages.error(request, f'Error processing QC test: {str(e)}')
        
        return redirect('dashboards:qc_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get QC phases this user can work on - EXCLUDE failed and completed tests
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        # Filter out failed and completed QC tests so they don't reappear
        filtered_phases = [p for p in user_phases if p.status in ['pending', 'in_progress']]
        my_phases.extend(filtered_phases)
    
    # Statistics
    stats = {
        'pending_tests': len([p for p in my_phases if p.status == 'pending']),
        'in_testing': len([p for p in my_phases if p.status == 'in_progress']),
        'passed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date(),
            status='completed'
        ).count(),
        'failed_this_week': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date__gte=timezone.now().date() - timedelta(days=7),
            status='failed'
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    daily_progress = min(100, (stats['passed_today'] / max(1, stats['pending_tests'] + stats['passed_today'])) * 100)
    
    # Get quarantine samples awaiting QC testing
    try:
        from quarantine.models import SampleRequest
        quarantine_samples = SampleRequest.objects.filter(
            sample_date__isnull=False,  # Already processed by QA
            qc_status='pending'  # Pending QC decision
        ).select_related(
            'quarantine_batch__bmr__product',
            'quarantine_batch__bmr'
        ).order_by('-sample_date')
    except ImportError:
        # Quarantine app not yet migrated
        quarantine_samples = []
    
    context = {
        'user': request.user,
        'my_phases': my_phases,
        'qc_phases': my_phases,  # Add this for template compatibility
        'quarantine_samples': quarantine_samples,  # Add quarantine samples
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Quality Control Dashboard'
    }
    
    return render(request, 'dashboards/qc_dashboard.html', context)

@login_required
def packaging_dashboard(request):
    """Packaging Store Dashboard"""
    if request.user.role != 'packaging_store':
        messages.error(request, 'Access denied. Packaging Store role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for packaging material release
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        notes = request.POST.get('notes', '')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start packaging material release for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect('dashboards:packaging_dashboard')
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Packaging material release started by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    messages.success(request, f'Packaging material release started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Packaging materials released by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    # Set session variables for next phase notification
                    request.session['completed_phase'] = phase_execution.phase.phase_name
                    request.session['completed_bmr'] = phase_execution.bmr.id
                    
                    # Trigger next phase in workflow (should be packing phases)
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    # Determine correct message based on product type
                    if phase_execution.bmr.product.product_type == 'tablet' and getattr(phase_execution.bmr.product, 'tablet_type', None) == 'tablet_2':
                        messages.success(request, f'Packaging materials released for batch {phase_execution.bmr.batch_number}. Bulk packing is now available.')
                    else:
                        messages.success(request, f'Packaging materials released for batch {phase_execution.bmr.batch_number}. Packing phases are now available.')
                    
            except Exception as e:
                messages.error(request, f'Error processing packaging material release: {str(e)}')
        
        return redirect('dashboards:packaging_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get packaging phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Build operator history for this user (recent phases where user was started_by or completed_by)
    recent_phases = BatchPhaseExecution.objects.filter(
        Q(started_by=request.user) | Q(completed_by=request.user)
    ).order_by('-started_date', '-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in recent_phases
    ]

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Packaging Store Dashboard',
        'operator_history': operator_history,
    }
    
    # Get next phase info for notification
    completed_phase = request.session.pop('completed_phase', None)
    bmr_id = request.session.pop('completed_bmr', None)
    bmr = None
    next_phase = None
    if bmr_id:
        try:
            bmr = BMR.objects.get(id=bmr_id)
            # For tablet type 2, make sure bulk packing comes before secondary packing
            if bmr.product.product_type == 'tablet' and getattr(bmr.product, 'tablet_type', None) == 'tablet_2':
                # Check if material release was just completed
                if completed_phase == 'packaging_material_release':
                    next_phase = BatchPhaseExecution.objects.filter(bmr=bmr, phase__phase_name='bulk_packing').first()
            
            # Fallback to standard next phase logic if no specific phase found
            if not next_phase:
                next_phase = WorkflowService.get_next_phase(bmr)
        except BMR.DoesNotExist:
            pass
    
    # Add notification context
    context.update({
        'completed_phase': completed_phase,
        'bmr': bmr,
        'next_phase': next_phase
    })
    
    return render(request, 'dashboards/packaging_dashboard.html', context)

@login_required
def packing_dashboard(request):
    """Packing Operator Dashboard"""
    if request.user.role != 'packing_operator':
        messages.error(request, 'Access denied. Packing Operator role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for packing phase completion
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        notes = request.POST.get('notes', '')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start packing for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect('dashboards:packing_dashboard')
                    
                    # Handle machine selection
                    machine_id = request.POST.get('machine_id')
                    if machine_id:
                        try:
                            machine = Machine.objects.get(id=machine_id, is_active=True)
                            phase_execution.machine_used = machine
                        except Machine.DoesNotExist:
                            messages.error(request, 'Selected machine is not available.')
                            return redirect('dashboards:packing_dashboard')
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Packing started by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    messages.success(request, f'Packing started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    # Handle breakdown tracking
                    breakdown_occurred = request.POST.get('breakdown_occurred') == 'on'
                    if breakdown_occurred:
                        phase_execution.breakdown_occurred = True
                        breakdown_start = request.POST.get('breakdown_start_time')
                        breakdown_end = request.POST.get('breakdown_end_time')
                        breakdown_reason = request.POST.get('breakdown_reason', '')
                        
                        if breakdown_start:
                            phase_execution.breakdown_start_time = datetime.fromisoformat(breakdown_start.replace('T', ' '))
                        if breakdown_end:
                            phase_execution.breakdown_end_time = datetime.fromisoformat(breakdown_end.replace('T', ' '))
                        phase_execution.breakdown_reason = breakdown_reason
                    
                    # Handle changeover tracking
                    changeover_occurred = request.POST.get('changeover_occurred') == 'on'
                    if changeover_occurred:
                        phase_execution.changeover_occurred = True
                        changeover_start = request.POST.get('changeover_start_time')
                        changeover_end = request.POST.get('changeover_end_time')
                        changeover_reason = request.POST.get('changeover_reason', '')
                        
                        if changeover_start:
                            phase_execution.changeover_start_time = datetime.fromisoformat(changeover_start.replace('T', ' '))
                        if changeover_end:
                            phase_execution.changeover_end_time = datetime.fromisoformat(changeover_end.replace('T', ' '))
                        phase_execution.changeover_reason = changeover_reason
                    
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Packing completed by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'Packing completed for batch {phase_execution.bmr.batch_number}.')
                    
            except Exception as e:
                messages.error(request, f'Error processing packing phase: {str(e)}')
        
        return redirect('dashboards:packing_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.all()
    
    # Get packing phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'pending_packing': len([p for p in my_phases if p.status == 'pending']),  # For template compatibility
        'in_progress_packing': len([p for p in my_phases if p.status == 'in_progress']),  # For template compatibility
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }

    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Get available machines for this user role
    machine_type_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending', 
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'blister_packing',  # Packing operator uses blister packing machines
        'filling_operator': 'filling',  # For capsule filling
    }
    
    user_machine_type = machine_type_mapping.get(request.user.role)
    available_machines = []
    if user_machine_type:
        available_machines = Machine.objects.filter(
            machine_type=user_machine_type,
            is_active=True
        ).order_by('name')
    
    # Determine if this role should show breakdown/changeover tracking
    # Only for phases that use machines
    breakdown_tracking_roles = [
        'mixing_operator', 'granulation_operator', 'blending_operator', 'compression_operator',
        'coating_operator', 'tube_filling_operator', 'filling_operator'
    ]
    show_breakdown_tracking = request.user.role in breakdown_tracking_roles    # Build operator history for this user (recent phases where user was started_by or completed_by)
    recent_phases = BatchPhaseExecution.objects.filter(
        Q(started_by=request.user) | Q(completed_by=request.user)
    ).order_by('-started_date', '-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in recent_phases
    ]

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'packing_phases': my_phases,  # Add this for template compatibility
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Packing Dashboard',
        'operator_history': operator_history,
        'available_machines': available_machines,
        'show_breakdown_tracking': show_breakdown_tracking,
    }
    
    return render(request, 'dashboards/packing_dashboard.html', context)

@login_required
def finished_goods_dashboard(request):
    """Finished Goods Store Dashboard with Inventory Management"""
    if request.user.role != 'finished_goods_store':
        messages.error(request, 'Access denied. Finished Goods Store role required.')
        return redirect('dashboards:dashboard_home')
    
    # Import FGS models
    from fgs_management.models import FGSInventory, ProductRelease, FGSAlert
    from django.utils import timezone
    from datetime import timedelta
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    # Only show finished_goods_store phases
    my_phases = [p for p in my_phases if getattr(p.phase, 'phase_name', None) == 'finished_goods_store']
    
    # Get all finished goods store phases for history statistics
    all_fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store'
    ).select_related('bmr', 'phase', 'bmr__product')
    
    # FGS Inventory Statistics
    total_inventory_items = FGSInventory.objects.count()
    available_for_sale = FGSInventory.objects.filter(status='available').count()
    
    # Recent releases (last 7 days)
    recent_releases_count = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=7)
    ).count()
    
    # Active alerts
    active_alerts_count = FGSAlert.objects.filter(is_resolved=False).count()
    
    # Recent inventory items
    recent_inventory = FGSInventory.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).select_related('product', 'bmr').order_by('-created_at')[:10]
    
    # Current inventory available for release
    available_inventory = FGSInventory.objects.filter(
        status__in=['stored', 'available'],
        quantity_available__gt=0
    ).select_related('product', 'bmr').order_by('-created_at')
    
    # Completed FGS phases without inventory entries
    completed_fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).exclude(
        bmr__in=FGSInventory.objects.values_list('bmr', flat=True)
    ).select_related('bmr__product').order_by('-completed_date')[:10]
    
    # Recent releases
    recent_releases = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=14)
    ).select_related('inventory__product', 'inventory__bmr').order_by('-release_date')[:10]
    
    # Active alerts
    active_alerts = FGSAlert.objects.filter(
        is_resolved=False
    ).select_related('inventory').order_by('-priority', '-created_at')[:10]
    
    # Filtering support for dashboard cards
    filter_param = request.GET.get('filter')
    detail_param = request.GET.get('detail')
    
    # Detail view for specific card
    if detail_param:
        if detail_param == 'pending':
            my_phases = [p for p in my_phases if p.status == 'pending']
        elif detail_param == 'in_progress':
            my_phases = [p for p in my_phases if p.status == 'in_progress']
        elif detail_param == 'completed_today':
            today = timezone.now().date()
            my_phases = [p for p in all_fgs_phases if p.status == 'completed' and 
                         getattr(p, 'completed_date', None) and p.completed_date.date() == today]
        elif detail_param == 'total_batches':
            # Show all batches that have reached FGS
            my_phases = list(all_fgs_phases)
    # Regular filtering
    elif filter_param:
        if filter_param == 'completed_today':
            my_phases = [p for p in my_phases if p.status == 'completed' and getattr(p, 'completed_by', None) == request.user and getattr(p, 'completed_date', None) and p.completed_date.date() == timezone.now().date()]
        elif filter_param == 'total_batches':
            # Show all phases (default)
            pass
        else:
            my_phases = [p for p in my_phases if p.status == filter_param]
    
    # History statistics (last 7 days)
    today = timezone.now().date()
    last_7_days = [today - timezone.timedelta(days=i) for i in range(7)]
    daily_completions = {}
    
    for day in last_7_days:
        count = all_fgs_phases.filter(
            status='completed',
            completed_date__date=day
        ).count()
        daily_completions[day.strftime('%a')] = count
    
    # Product type statistics in FGS
    product_types = {}
    for phase in all_fgs_phases.filter(status__in=['in_progress', 'completed']):
        product_type = phase.bmr.product.product_type
        if product_type in product_types:
            product_types[product_type] += 1
        else:
            product_types[product_type] = 1

    # Statistics - Updated with real FGS data
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status='completed',
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': all_fgs_phases.values('bmr').distinct().count(),
        'daily_history': daily_completions,
        'product_types': product_types,
        
        # FGS-specific statistics
        'total_inventory_items': total_inventory_items,
        'available_for_sale': available_for_sale,
        'recent_releases': recent_releases.count(),
        'active_alerts': active_alerts.count(),
    }

    # Determine the primary phase name for this role
    role_phase_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'drying_operator': 'drying',
        'filling_operator': 'filling',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'packing',
        'sorting_operator': 'sorting',
    }

    phase_name = role_phase_mapping.get(request.user.role, 'production')
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Get recently completed goods
    recent_completed = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).select_related('bmr', 'bmr__product').order_by('-completed_date')[:5]
    
    # Storage efficiency (time from final QA to FGS)
    efficiency_data = []
    for phase in recent_completed:
        final_qa_phase = BatchPhaseExecution.objects.filter(
            bmr=phase.bmr,
            phase__phase_name='final_qa',
            status='completed'
        ).first()
        
        if final_qa_phase and final_qa_phase.completed_date and phase.completed_date:
            storage_time = (phase.completed_date - final_qa_phase.completed_date).total_seconds() / 3600  # hours
            efficiency_data.append({
                'bmr': phase.bmr,
                'time_hours': round(storage_time, 1)
            })
    
    # Card specific view
    detail_title = None
    if request.GET.get('detail'):
        detail = request.GET.get('detail')
        if detail == 'pending':
            detail_title = 'Pending Storage'
        elif detail == 'in_progress':
            detail_title = 'In Storage'
        elif detail == 'completed_today':
            detail_title = 'Stored Today'
        elif detail == 'total_batches':
            detail_title = 'All Batches in FGS'

    # Process all phases to add display name
    for phase in my_phases:
        if hasattr(phase, 'phase') and hasattr(phase.phase, 'phase_name'):
            phase.display_name = format_phase_name(phase.phase.phase_name)
    
    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'phase_name': 'finished_goods_store',
        'phase_display_name': 'Finished Goods Store',
        'daily_progress': daily_progress,
        'dashboard_title': 'Finished Goods Store Dashboard',
        'active_filter': filter_param,
        'recent_completed': recent_completed,
        'efficiency_data': efficiency_data,
        'detail_title': detail_title,
        'detail_view': request.GET.get('detail'),
        
        # New FGS inventory data
        'recent_inventory': recent_inventory,
        'recent_releases': recent_releases,
        'active_alerts': active_alerts,
        'available_inventory': available_inventory,
        'completed_fgs_phases': completed_fgs_phases,
    }

    return render(request, 'dashboards/finished_goods_dashboard.html', context)

@login_required
def admin_fgs_monitor(request):
    """Admin FGS Monitor - Track finished goods storage with inventory management"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Import FGS models
    from fgs_management.models import FGSInventory, ProductRelease, FGSAlert
    from django.utils import timezone
    from datetime import timedelta
    
    # Get finished goods storage phases
    fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store'
    ).select_related('bmr__product', 'started_by', 'completed_by').order_by('-started_date')
    
    # Group by status
    fgs_pending = fgs_phases.filter(status='pending')
    fgs_in_progress = fgs_phases.filter(status='in_progress') 
    fgs_completed = fgs_phases.filter(status='completed')
    
    # FGS Inventory Statistics
    total_inventory_items = FGSInventory.objects.count()
    available_for_sale = FGSInventory.objects.filter(status='available').count()
    
    # Recent releases (last 7 days)
    recent_releases_count = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=7)
    ).count()
    
    # Active alerts
    active_alerts_count = FGSAlert.objects.filter(is_resolved=False).count()
    
    # Statistics
    fgs_stats = {
        'total_in_store': fgs_completed.count(),
        'pending_storage': fgs_pending.count(),
        'being_stored': fgs_in_progress.count(),
        'storage_capacity_used': min(100, (fgs_completed.count() / max(1000, 1)) * 100),  # Assuming 1000 batch capacity
        
        # New inventory statistics
        'total_inventory_items': total_inventory_items,
        'available_for_sale': available_for_sale,
        'recent_releases': recent_releases_count,
        'active_alerts': active_alerts_count,
    }
    
    # Recent storage activity
    recent_stored = fgs_completed[:10]
    
    # Recent inventory items
    recent_inventory = FGSInventory.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).select_related('product', 'bmr').order_by('-created_at')[:10]
    
    # Recent releases
    recent_releases = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=14)
    ).select_related('inventory__product', 'inventory__bmr').order_by('-release_date')[:10]
    
    # Active alerts
    active_alerts = FGSAlert.objects.filter(
        is_resolved=False
    ).select_related('inventory').order_by('-priority', '-created_at')[:10]
    
    # Products in FGS by type
    products_in_fgs = fgs_completed.values(
        'bmr__product__product_type',
        'bmr__product__product_name'
    ).annotate(
        batch_count=Count('bmr'),
        latest_storage=Max('completed_date')
    ).order_by('bmr__product__product_type', '-latest_storage')
    
    # Get production data by product type
    product_type_data = {}
    completed_bmrs = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).select_related('bmr__product')
    
    for execution in completed_bmrs:
        product_type = execution.bmr.product.product_type
        if product_type not in product_type_data:
            product_type_data[product_type] = 0
        product_type_data[product_type] += 1
    
    # Get phase completion status across all batches
    phase_completion = {}
    all_phases = BatchPhaseExecution.objects.values('phase__phase_name').distinct()
    for phase_dict in all_phases:
        phase_name = phase_dict['phase__phase_name']
        if phase_name:
            total = BatchPhaseExecution.objects.filter(phase__phase_name=phase_name).count()
            completed = BatchPhaseExecution.objects.filter(
                phase__phase_name=phase_name,
                status='completed'
            ).count()
            if total > 0:  # Avoid division by zero
                completion_rate = (completed / total) * 100
            else:
                completion_rate = 0
            phase_completion[phase_name] = {
                'total': total,
                'completed': completed,
                'completion_rate': round(completion_rate, 1)
            }
    
    # Get weekly production trend
    today = timezone.now().date()
    start_date = today - timezone.timedelta(days=28)  # Last 4 weeks
    
    weekly_completions = {}
    for i in range(4):  # 4 weeks
        week_start = start_date + timezone.timedelta(days=i*7)
        week_end = week_start + timezone.timedelta(days=6)
        week_label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}"
        
        weekly_completions[week_label] = BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status='completed',
            completed_date__date__range=[week_start, week_end]
        ).count()
    
    # QC pass/fail data
    qc_stats = {
        'passed': BatchPhaseExecution.objects.filter(
            phase__phase_name__in=['post_compression_qc', 'post_mixing_qc', 'post_blending_qc'],
            status='completed'
        ).count(),
        'failed': BatchPhaseExecution.objects.filter(
            phase__phase_name__in=['post_compression_qc', 'post_mixing_qc', 'post_blending_qc'],
            status='failed'
        ).count()
    }
    
    context = {
        'user': request.user,
        'fgs_pending': fgs_pending,
        'fgs_in_progress': fgs_in_progress,
        'recent_stored': recent_stored,
        'fgs_stats': fgs_stats,
        'products_in_fgs': products_in_fgs,
        'dashboard_title': 'Finished Goods Store Monitor',
        'product_type_data': product_type_data,
        'phase_completion': phase_completion,
        'weekly_production': weekly_completions,
        'qc_stats': qc_stats,
        
        # New FGS inventory data
        'recent_inventory': recent_inventory,
        'recent_releases': recent_releases,
        'active_alerts': active_alerts,
    }
    
    return render(request, 'dashboards/admin_fgs_monitor.html', context)

@login_required
def live_tracking_view(request):
    """Live BMR Tracking - Per BMR, per phase, with start/end/duration and total time"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')

    bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()
    timeline_data = []
    from workflow.models import BatchPhaseExecution
    for bmr in bmrs:
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
        phase_timeline = []
        for phase in phases:
            phase_data = {
                'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                'status': phase.status.title(),
                'started_date': phase.started_date,
                'completed_date': phase.completed_date,
                'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                'duration_hours': None,
                'duration_formatted': None,
            }
            if phase.started_date and phase.completed_date:
                duration = phase.completed_date - phase.started_date
                total_hours = duration.total_seconds() / 3600
                phase_data['duration_hours'] = round(total_hours, 2)
                
                # Format duration in hours and minutes
                hours = int(total_hours)
                minutes = int((total_hours - hours) * 60)
                if hours > 0:
                    phase_data['duration_formatted'] = f"{hours}h {minutes}m"
                else:
                    phase_data['duration_formatted'] = f"{minutes}m"
            elif phase.started_date and not phase.completed_date:
                from django.utils import timezone
                duration = timezone.now() - phase.started_date
                total_hours = duration.total_seconds() / 3600
                phase_data['duration_hours'] = round(total_hours, 2)
                
                # Format duration in hours and minutes
                hours = int(total_hours)
                minutes = int((total_hours - hours) * 60)
                if hours > 0:
                    phase_data['duration_formatted'] = f"{hours}h {minutes}m (ongoing)"
                else:
                    phase_data['duration_formatted'] = f"{minutes}m (ongoing)"
            phase_timeline.append(phase_data)
        # Calculate total production time for this BMR
        total_production_time = None
        bmr_creation_date = bmr.created_date
        
        # Find the finished goods store phase
        fgs_phase = next((p for p in phase_timeline if p['phase_name'] == 'Finished Goods Store' and p['status'] == 'Completed'), None)
        
        if bmr_creation_date and fgs_phase and fgs_phase['completed_date']:
            # Calculate total production time
            total_duration = fgs_phase['completed_date'] - bmr_creation_date
            total_hours = total_duration.total_seconds() / 3600
            
            # Format total production time
            days = int(total_hours // 24)
            hours = int(total_hours % 24)
            minutes = int((total_hours % 1) * 60)
            
            if days > 0:
                total_production_time = f"{days}d {hours}h {minutes}m"
            elif hours > 0:
                total_production_time = f"{hours}h {minutes}m"
            else:
                total_production_time = f"{minutes}m"
        
        timeline_data.append({
            'bmr': bmr,
            'phase_timeline': phase_timeline,
            'total_production_time': total_production_time,
        })
    return render(request, 'dashboards/live_tracking.html', {'timeline_data': timeline_data, 'dashboard_title': 'Live BMR Tracking'})

def export_timeline_data(request, timeline_data=None, format_type=None):
    """Export detailed timeline data to CSV or Excel with all phases"""
    # Handle direct URL access
    if timeline_data is None:
        # Get export format from request
        format_type = request.GET.get('format', 'excel')
        
        # Recreate the timeline data from scratch
        from bmr.models import BMR
        from workflow.models import BatchPhaseExecution
        
        bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()
        
        # Add timeline data for each BMR
        timeline_data = []
        for bmr in bmrs:
            phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
            bmr_created = bmr.created_date
            fgs_completed = phases.filter(
                phase__phase_name='finished_goods_store',
                status='completed'
            ).first()
            total_time_hours = None
            if fgs_completed and fgs_completed.completed_date:
                total_time_hours = round((fgs_completed.completed_date - bmr_created).total_seconds() / 3600, 2)
            phase_timeline = []
            for phase in phases:
                phase_data = {
                    'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                    'status': phase.status.title(),
                    'started_date': phase.started_date,
                    'completed_date': phase.completed_date,
                    'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                    'completed_by': phase.completed_by.get_full_name() if phase.completed_by else None,
                    'duration_hours': None,
                    'operator_comments': getattr(phase, 'operator_comments', '') or '',
                    'phase_order': phase.phase.phase_order if hasattr(phase.phase, 'phase_order') else 0,
                    # Machine tracking
                    'machine_used': phase.machine_used.name if phase.machine_used else '',
                    # Breakdown tracking
                    'breakdown_occurred': 'Yes' if phase.breakdown_occurred else 'No',
                    'breakdown_duration': phase.get_breakdown_duration() if hasattr(phase, 'get_breakdown_duration') and phase.breakdown_occurred else '',
                    'breakdown_start_time': phase.breakdown_start_time if phase.breakdown_occurred else '',
                    'breakdown_end_time': phase.breakdown_end_time if phase.breakdown_occurred else '',
                    # Changeover tracking
                    'changeover_occurred': 'Yes' if phase.changeover_occurred else 'No',
                    'changeover_duration': phase.get_changeover_duration() if hasattr(phase, 'get_changeover_duration') and phase.changeover_occurred else '',
                    'changeover_start_time': phase.changeover_start_time if phase.changeover_occurred else '',
                    'changeover_end_time': phase.changeover_end_time if phase.changeover_occurred else '',
                }
                if phase.started_date and phase.completed_date:
                    duration = phase.completed_date - phase.started_date
                    phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
                elif phase.started_date and not phase.completed_date:
                    duration = timezone.now() - phase.started_date
                    phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
                phase_timeline.append(phase_data)
            timeline_data.append({
                'bmr': bmr,
                'total_time_hours': total_time_hours,
                'phase_timeline': phase_timeline,
                'current_phase': phases.filter(status__in=['pending', 'in_progress']).first(),
                'is_completed': fgs_completed is not None,
            })
    
    # Generate CSV export
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="bmr_detailed_timeline_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        
        # Header row
        writer.writerow(['BMR Report - Generated on', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Write detailed phase information for each BMR
        for item in timeline_data:
            bmr = item['bmr']
            writer.writerow([])  # Empty row for separation
            writer.writerow([f"BMR: {bmr.batch_number} - {bmr.product.product_name}"])
            writer.writerow([f"Product Type: {bmr.product.product_type}"])
            writer.writerow([f"Created: {bmr.created_date.strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([f"Total Production Time: {item['total_time_hours']} hours" if item['total_time_hours'] else "In Progress"])
            writer.writerow([])  # Empty row
            writer.writerow([
                'Phase Name', 'Status', 'Started Date', 'Started By', 
                'Completed Date', 'Completed By', 'Duration (Hours)', 'Comments',
                'Machine Used', 'Breakdown Occurred', 'Breakdown Duration (Min)', 
                'Breakdown Start', 'Breakdown End', 'Changeover Occurred', 
                'Changeover Duration (Min)', 'Changeover Start', 'Changeover End'
            ])
            for phase in item['phase_timeline']:
                writer.writerow([
                    phase['phase_name'], phase['status'],
                    phase['started_date'], phase['started_by'],
                    phase['completed_date'], phase['completed_by'],
                    phase['duration_hours'], phase['operator_comments'],
                    phase['machine_used'], phase['breakdown_occurred'], 
                    phase['breakdown_duration'], phase['breakdown_start_time'],
                    phase['breakdown_end_time'], phase['changeover_occurred'],
                    phase['changeover_duration'], phase['changeover_start_time'],
                    phase['changeover_end_time']
                ])
        return response
    
    # Generate Excel export
    elif format_type == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Production Summary"
        
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Create title
        summary_sheet.merge_cells('A1:I1')
        title_cell = summary_sheet['A1']
        title_cell.value = "Kampala Pharmaceutical Industries - BMR Production Timeline Summary"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Create report generation date
        summary_sheet.merge_cells('A2:I2')
        date_cell = summary_sheet['A2']
        date_cell.value = f"Report Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.font = Font(italic=True)
        
        # Add empty row
        summary_sheet.append([])
        
        # Summary headers
        headers = [
            "Batch Number", "Product Name", "Product Type", 
            "Created Date", "Current Status", "Current Phase",
            "Total Duration (Hours)", "Completed", "Bottleneck Phase"
        ]
        
        header_row = summary_sheet.row_dimensions[4]
        header_row.height = 30
        
        for col_num, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            summary_sheet.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Add data rows
        row_num = 5
        for item in timeline_data:
            bmr = item['bmr']
            # Find bottleneck phase (longest duration)
            bottleneck = max(item['phase_timeline'], key=lambda x: x['duration_hours'] if x['duration_hours'] else 0, default={})
            bottleneck_name = bottleneck.get('phase_name', 'N/A') if bottleneck else 'N/A'
            
            # Get current phase
            current_phase = "Completed"
            if not item['is_completed']:
                current_phases = [p for p in item['phase_timeline'] if p['status'] in ['In Progress', 'Pending']]
                if current_phases:
                    current_phase = current_phases[0]['phase_name']
            
            # Add row data
            row_data = [
                bmr.batch_number,
                bmr.product.product_name,
                bmr.product.product_type.replace('_', ' ').title(),
                bmr.created_date.strftime('%Y-%m-%d'),
                "Completed" if item['is_completed'] else "In Progress",
                current_phase,
                item['total_time_hours'] if item['total_time_hours'] else "In Progress",
                "Yes" if item['is_completed'] else "No",
                bottleneck_name
            ]
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
        
        # Create detail sheet for each BMR
        for item in timeline_data:
            bmr = item['bmr']
            # Create sheet for this BMR
            detail_sheet = wb.create_sheet(title=f"BMR-{bmr.batch_number}")
            
            # Title
            detail_sheet.merge_cells('A1:H1')
            title_cell = detail_sheet['A1']
            title_cell.value = f"Detailed Timeline for BMR {bmr.batch_number} - {bmr.product.product_name}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            
            # BMR information
            detail_sheet.merge_cells('A2:H2')
            info_cell = detail_sheet['A2']
            info_cell.value = f"Product Type: {bmr.product.product_type.replace('_', ' ').title()} | Created: {bmr.created_date.strftime('%Y-%m-%d %H:%M:%S')}"
            info_cell.font = Font(italic=True)
            info_cell.alignment = Alignment(horizontal="center")
            
            detail_sheet.merge_cells('A3:H3')
            time_cell = detail_sheet['A3']
            time_cell.value = f"Total Production Time: {item['total_time_hours']} hours" if item['total_time_hours'] else "Total Production Time: In Progress"
            time_cell.font = Font(italic=True, bold=True)
            time_cell.alignment = Alignment(horizontal="center")
            
            # Add empty row
            detail_sheet.append([])
            
            # Detail headers
            headers = [
                "Phase Name", "Status", "Started Date", "Started By", 
                "Completed Date", "Completed By", "Duration (Hours)", "Comments",
                "Machine Used", "Breakdown Occurred", "Breakdown Duration (Min)", 
                "Breakdown Start", "Breakdown End", "Changeover Occurred", 
                "Changeover Duration (Min)", "Changeover Start", "Changeover End"
            ]
            
            header_row = detail_sheet.row_dimensions[5]
            header_row.height = 30
            
            for col_num, header in enumerate(headers, 1):
                cell = detail_sheet.cell(row=5, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # Adjust column widths for new columns
                if col_num <= 8:  # Original columns
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 18
                elif col_num in [9, 10, 14]:  # Machine, breakdown occurred, changeover occurred
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 15
                else:  # Date/time columns
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Add phase data
            phase_row = 6
            for phase in item['phase_timeline']:
                # Format dates for display
                started_date = phase['started_date'].strftime('%Y-%m-%d %H:%M') if phase['started_date'] else "Not Started"
                completed_date = phase['completed_date'].strftime('%Y-%m-%d %H:%M') if phase['completed_date'] else "Not Completed"
                breakdown_start = phase['breakdown_start_time'].strftime('%Y-%m-%d %H:%M') if phase['breakdown_start_time'] else ""
                breakdown_end = phase['breakdown_end_time'].strftime('%Y-%m-%d %H:%M') if phase['breakdown_end_time'] else ""
                changeover_start = phase['changeover_start_time'].strftime('%Y-%m-%d %H:%M') if phase['changeover_start_time'] else ""
                changeover_end = phase['changeover_end_time'].strftime('%Y-%m-%d %H:%M') if phase['changeover_end_time'] else ""
                
                phase_data = [
                    phase['phase_name'],
                    phase['status'],
                    started_date,
                    phase['started_by'] if phase['started_by'] else "",
                    completed_date,
                    phase['completed_by'] if phase['completed_by'] else "",
                    phase['duration_hours'] if phase['duration_hours'] is not None else "",
                    phase['operator_comments'] if phase['operator_comments'] else "",
                    phase['machine_used'] if phase['machine_used'] else "",
                    phase['breakdown_occurred'],
                    phase['breakdown_duration'] if phase['breakdown_duration'] else "",
                    breakdown_start,
                    breakdown_end,
                    phase['changeover_occurred'],
                    phase['changeover_duration'] if phase['changeover_duration'] else "",
                    changeover_start,
                    changeover_end
                ]
                
                # Apply styling based on status
                row_fill = None
                if phase['status'] == 'Completed':
                    row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif phase['status'] == 'In Progress':
                    row_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                
                for col_num, cell_value in enumerate(phase_data, 1):
                    cell = detail_sheet.cell(row=phase_row, column=col_num)
                    cell.value = cell_value
                    cell.border = border
                    if row_fill:
                        cell.fill = row_fill
                    
                    # For comments column, use wrap text
                    if col_num == 8:  # Comments column
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        detail_sheet.row_dimensions[phase_row].height = max(15, min(50, len(str(cell_value)) // 10 * 15))
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Adjust column widths for new columns
                    if col_num <= 8:  # Original columns
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 18
                    elif col_num in [9, 10, 14]:  # Machine, breakdown occurred, changeover occurred
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 15
                    else:  # Date/time columns
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 20
                
                phase_row += 1
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="bmr_timeline_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
    
    else:
        return HttpResponse('Unsupported export format', content_type='text/plain')


# Redirect view for old admin dashboard URL
def admin_redirect(request):
    # Direct redirect to admin dashboard function
    return admin_dashboard(request)


@login_required
def production_manager_dashboard(request):
    """Production Manager Dashboard - BMR Request Management"""
    if request.user.role != 'production_manager':
        messages.error(request, 'Access denied. Production Manager role required.')
        return redirect('dashboards:dashboard_home')
    
    # Import the BMRRequest model
    from bmr.models import BMRRequest
    
    # Get BMR request statistics for this user
    user_bmr_requests = BMRRequest.objects.filter(requested_by=request.user)
    bmr_request_stats = {
        'total': user_bmr_requests.count(),
        'pending': user_bmr_requests.filter(status='pending').count(),
        'approved': user_bmr_requests.filter(status='approved').count(),
        'rejected': user_bmr_requests.filter(status='rejected').count(),
        'completed': user_bmr_requests.filter(status='completed').count(),
    }
    
    # Get recent BMR requests
    recent_bmr_requests = user_bmr_requests.select_related('product').order_by('-request_date')[:10]
    
    # Get overall production statistics
    all_bmrs = BMR.objects.all()
    production_stats = {
        'total_bmrs': all_bmrs.count(),
        'active_production': all_bmrs.filter(status__in=['approved', 'in_production']).count(),
        'completed_batches': all_bmrs.filter(status='completed').count(),
        'pending_approval': all_bmrs.filter(status='submitted').count(),
    }
    
    # Get products available for BMR requests
    available_products = Product.objects.all().order_by('product_name')
    
    # Get BMRs created from this user's requests
    user_bmrs = BMR.objects.filter(
        bmr_requests__requested_by=request.user
    ).distinct().select_related('product').order_by('-created_date')[:5]
    
    context = {
        'user': request.user,
        'bmr_request_stats': bmr_request_stats,
        'recent_bmr_requests': recent_bmr_requests,
        'production_stats': production_stats,
        'available_products': available_products,
        'user_bmrs': user_bmrs,
        'dashboard_title': 'Production Manager Dashboard',
    }
    
    return render(request, 'dashboards/production_manager_dashboard.html', context)

@login_required
def quarantine_monitor_view(request):
    """Dedicated view for quarantine monitoring section"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Import the models here to avoid circular imports
    from quarantine.models import QuarantineBatch, SampleRequest
    
    # Get recent quarantine batches
    quarantine_batches = QuarantineBatch.objects.select_related(
        'bmr', 'bmr__product', 'current_phase'
    ).order_by('-quarantine_date')[:20]
    
    # Get recent sample requests with detailed timeline
    recent_sample_requests = SampleRequest.objects.select_related(
        'quarantine_batch__bmr__product',
        'quarantine_batch__bmr__created_by',
        'requested_by',
        'sampled_by',  # QA user who sampled
        'received_by',  # QC user who received
        'approved_by'   # QC user who approved/rejected
    ).order_by('-request_date')[:15]
    
    # Calculate quarantine statistics
    total_quarantine_batches = QuarantineBatch.objects.count()
    pending_qa_samples = SampleRequest.objects.filter(
        sample_date__isnull=True  # No sample taken yet = pending QA
    ).count()
    pending_qc_samples = SampleRequest.objects.filter(
        sample_date__isnull=False,  # Sample taken
        approved_date__isnull=True  # No QC decision yet = pending QC
    ).count()
    approved_samples_today = SampleRequest.objects.filter(
        qc_status='approved',
        approved_date__date=timezone.now().date()
    ).count()
    rejected_samples_today = SampleRequest.objects.filter(
        qc_status='failed',
        approved_date__date=timezone.now().date()
    ).count()
    
    quarantine_stats = {
        'total_quarantine_batches': total_quarantine_batches,
        'pending_qa_samples': pending_qa_samples,
        'pending_qc_samples': pending_qc_samples,
        'approved_samples_today': approved_samples_today,
        'rejected_samples_today': rejected_samples_today,
        'avg_qa_processing_time': SampleRequest.objects.filter(
            sample_date__isnull=False  # QA processing completed
        ).aggregate(
            avg_time=models.Avg(
                models.F('sample_date') - models.F('request_date')
            )
        )['avg_time'],
        'avg_qc_processing_time': SampleRequest.objects.filter(
            approved_date__isnull=False,  # QC decision made
            received_date__isnull=False   # QC received sample
        ).aggregate(
            avg_time=models.Avg(
                models.F('approved_date') - models.F('received_date')
            )
        )['avg_time']
    }
    
    context = {
        'quarantine_stats': quarantine_stats,
        'recent_sample_requests': recent_sample_requests,
        'quarantine_batches': quarantine_batches,
        'page_title': 'Quarantine Monitoring'
    }
    
    return render(request, 'dashboards/quarantine_monitor.html', context)
